from openpyxl import Workbook
import openpyxl
from time import sleep 
import os
import random
program_path = os.path.dirname(os.path.abspath(__file__))
all_files = []
for (dirpath, dirnames, filenames) in os.walk(program_path):
    all_files.extend(filenames)
    break
wb = Workbook()
with open(program_path+"\\"+"settings.txt", 'r', encoding='utf-8') as settings:
    settingvalues = settings.readlines()

sheet = wb.active
mode = 0
# 1 = learn
# 2 = add
ui = None
path = None
chapter = None
loaded_workbook = False
savetick = 1
time_to_save = 0
correct_words = 0
wrong_words = 0
words = []
words_found = 0
just_commanded = False
prev_word = ""
ask_word_again = False

standard_savetick = 1
split_character = "  "
def gsay(colour1, colour2, text):
    #takes the first colour, the second colour and text and outputs a text with a gradient between the two colours.
    colour1 = list(int(colour1[i:i+2], 16) for i in (0, 2, 4))
    colour2 = list(int(colour2[i:i+2], 16) for i in (0, 2, 4))
    length = len(text)
    textlist = list(text)
    colourdiff = ([colour2[x] - colour1[x] for x in range(len(colour1))])
    colourstep = [int(colourdiff[x]/length) for x in range (len(colour1))]
    for i in range(0,length):
        print(str(f"\u001b[38;2;{colour1[0]+(colourstep[0]*i)};{colour1[1]+(colourstep[1]*i)};{colour1[2]+(colourstep[2]*i)}m"+textlist[i]), end="")
    print("\033[0m")

def say(text):
    texttosay = "\u001b[38;2;126;126;143m"+text
    print(texttosay)
    print("\033[0m", end = "")

def warn(text: str):
    problem = "\u001b[38;2;176;48;92m"+text
    print(problem)
    print("\033[0m", end = "")

def completed():
    tosay = "✅"
    print(tosay)
    sleep(0.5)
    print("\033[1A", end="")
    print("\033[K", end="")

def saved():
    tosay = "💾"
    print(tosay)
    sleep(0.5)
    print(f"\033[1A", end="")
    print("\033[K", end="")

def create_difficult_sheet():
    if "difficult" in wb.sheetnames:
        return
    ws = wb.create_sheet("difficult")
    wb.save(path)

with open(program_path+"\\"+"settings.txt", 'w', encoding='utf-8') as settings:
    settings.write(str(standard_savetick)+"\n")
    settings.write(str(savetick))

name = "Trainer Classic"
gsay("473B78", "B0305C", name)
say("v1.4")
warn("UNSUPPORTED VERSION! MAJOR FILE-BREAKING BUGS!")
#say("use help for more info")
while True:
    #select mode
    while mode == 0:
        ui = input("\u001b[38;2;165;72;172m> ")
        print("\033[0m", end = "")

        if ui == "quit":
            break

        if ui == "add":
            if loaded_workbook == False:
                warn("No file opened! use /file.open <file>")
                continue
            if chapter == None:
                warn("No chapter opened! use chptr.open <chapter>")
                continue
            os.system('cls')
            #loads data saved at the top of the file
            ws = wb.active
            index = ws.max_row
            if index == 1:
                index = 0
            if savetick != int(settingvalues[1]):
                settingvalues[1] = savetick
                with open(program_path+"\\"+"settings.txt", 'w', encoding='utf-8') as settings:
                    settings.writelines(settingvalues)
            
            savetick = int(settingvalues[1])

            say(f"selected file {path}")
            say(f"selected chapter: {chapter}")
            say(f"syntax: <learning language_word>{split_character}<motherlanguage_word>")
            say("use /quit to exit")
            mode = 2

        if ui == "learn":
            if loaded_workbook == False:
                warn("No file opened! use /file.open <file>")
                continue
            if chapter == None:
                warn("No chapter opened! use chptr.open <chapter>")
                continue
            if ws.cell(row=2, column=1).value == None:
                warn("there are no words in the opened chapter/file!")
            os.system('cls')
            ws = wb.active
            #loads data saved in settings.txt
            if savetick != int(settingvalues[1]):
                settingvalues[1] = savetick
                with open(program_path+"\\"+"settings.txt", 'w', encoding='utf-8') as settings:
                    settings.writelines(settingvalues)
            
            savetick = int(settingvalues[1])
            say(f"selected file {path}")
            say(f"selected chapter: {chapter}")
            say("use /quit to exit")
            mode = 1
            
        if ui == "savetick":
            say(f"current savetick: {savetick}")
            savetick = int(ui.replace("/savetick ", ""))
            settingvalues[1] = savetick
            with open(program_path+"\\"+"settings.txt", 'w', encoding='utf-8') as settings:
                settings.writelines(settingvalues)

        if ui == "list":
            warn(f"selected file {path}")
            warn(f"selected chapter: {chapter}")
            if ws.cell(row=1, column=1).value == None:
                say("No words in chapter")
            else:
                for i in range(1, ws.max_row+1):
                    sleep(0.01)
                    say(str(ws.cell(row=i, column=1).value+split_character+str(ws.cell(row=i, column=2).value)))
        
        #open file
        if "file" in ui:
            #sets a workbook
            if ".open " in ui:
                #gets all files
                all_files = []
                for (dirpath, dirnames, filenames) in os.walk(program_path):
                    all_files.extend(filenames)
                    break
                file_to_open = ui.replace("file.open ", "")
                file = str(file_to_open + ".xlsx")
                if file in all_files:
                    path = str(program_path+"\\"+file)
                    wb = openpyxl.load_workbook(path)
                    all_created_sheets = wb.sheetnames
                    create_difficult_sheet()
                    loaded_workbook = True
                    completed()
                else:
                    option = input("\u001b[38;2;165;72;172mfile not found. create new(y/n)? ")
                    if option == "y":
                        file_to_create = file_to_open + ".xlsx"
                        ws =  wb.active
                        ws.title = file_to_open
                        path = str(program_path+"\\"+file_to_create)
                        wb.save(path)
                        wb = openpyxl.load_workbook(path)
                        create_difficult_sheet()
                        loaded_workbook = True
                        completed()
            #deletes a file
            if ".delete " in ui:
                file_to_delete = ui.replace("file.delete ","")
                path = str(program_path+"\\"+file_to_delete+".xlsx")
                option = input(f"\u001b[38;2;165;72;172mare you sure you want to delete \"{file_to_delete}\" (y/n)?")
                print("\033[0m", end = "")
                if option == "y":
                    if file_to_delete+".xlsx" in all_files:
                        os.remove(path)
                    else:
                        say(f"file \"{file_to_delete}\" not found. Perhaps you spelled the name wrong?")
                else:
                    completed()  
            #create new file
            if ".create " in ui:
                file_title = ui.replace("file.create ", "")
                file_to_create = file_title + ".xlsx"
                ws =  wb.active
                ws.title = file_title
                path = str(program_path+"\\"+file_to_create)
                wb.save(path)
                wb = openpyxl.load_workbook(path)
                create_difficult_sheet()
                loaded_workbook = True
                completed()

            if ".list" in ui:
                for i in range(0,len(all_files) - 1):
                    if ".xlsx" in all_files[i]:
                        say(all_files[i])

        #everything that starts with chapter
        if "chptr" in ui:
            if loaded_workbook == False:
                warn("No file opened! use /file.open <file>")
                continue
            #selects a chapter or creates a new one (sheet)
            if ".open " in ui:
                chapter = ui.replace("chptr.open ", "")
                all_created_sheets = wb.sheetnames
                if chapter in all_created_sheets:
                    wb.active = wb[chapter]
                    ws = wb[chapter]
                    completed()
                else:
                    option = input("\u001b[38;2;165;72;172mchapter not found. create new(y/n)? ")
                    if option == "y":
                        ws = wb.create_sheet(chapter)
                        if all_created_sheets[0] in all_files:
                            wb.remove(all_created_sheets[0])
                        wb.save(path)
                        wb.active = wb[chapter]
                        completed()
            #removes a chapter(sheet)
            if ".delete " in ui:
                all_created_sheets = wb.sheetnames
                chapter_to_delete = ui.replace("chptr.delete ", "")
                option = input(f"\u001b[38;2;165;72;172mare you sure you want to delete \"{chapter_to_delete}\" (y/n)?")
                print("\033[0m", end = "")
                if option == "y":
                    if chapter_to_delete in all_created_sheets:
                        wb.remove(wb[str(chapter_to_delete)])
                        wb.save(path)
                        completed()
                    else:
                        say(f"chapter \"{chapter_to_delete}\" not found. Perhaps you spelled the name wrong?")
                else:
                    completed()
            #creates a chapter (sheet)
            if ".create " in ui:
                all_created_sheets = wb.sheetnames
                chapter_to_create = ui.replace("chptr.create ", "")
                ws = wb.create_sheet(chapter_to_create)
                if all_created_sheets[0] in all_files:
                    wb.remove(all_created_sheets[0])
                wb.save(path)
                completed()

            if ".list" in ui:
                all_created_sheets = wb.sheetnames
                for i in range(0, len(all_created_sheets)):
                    say(all_created_sheets[i])

            if ".next" in ui:
                curr_idx = wb.sheetnames.index(chapter)
                if curr_idx + 1 < len(wb.sheetnames):
                    chapter = wb.sheetnames[curr_idx+1]
                    wb.active = wb[chapter]
                    ws = wb[chapter]
                    completed()

                else:
                    warn("no next chapter")

    #learn mode
    while mode == 1:
        while True:
            option = input("\u001b[38;2;234;169;77m"+"only repeat previously wrong words(y/n)? ")

            if option == "n":
                say("loading...")
                for i in range(1,ws.max_row+1):
                    words.append([ws.cell(row=i, column=2).value, ws.cell(row=i, column=1).value, i])
                break

            if option == "y":
                say("loading...")
                for i in range(1,ws.max_row+1):
                    if ws.cell(row=i, column= 3).value == "wrong" or ws.cell(row=i, column= 3).value == None:
                        words.append([ws.cell(row=i, column=2).value, ws.cell(row=i, column=1).value, i])

                if words == []:
                    warn("All words are correct!")
                    
                break
            
            else:
                warn("invalid input!")
        

        random.shuffle(words)
        index = 0
        increase_index = False
        difficult_words = []
        while True:
            if index % savetick == 0:
                wb.save(path)
                saved()
            if increase_index:
                if index < (len(words)-1):
                    index += 1
                else:
                    break

            increase_index = True

            mword = words[index][0]
            word = words[index][1]
            file_index = words[index][2]

            width = os.get_terminal_size()[0]

            if correct_words == 0:
                bar_correct = ""
            else:
                bar_correct = "█"*(round((correct_words/(len(words)) * width)))
            
            if  wrong_words == 0:
                bar_wrong = ""
            else:
                bar_wrong = "█" * (round((wrong_words/(len(words)) * width)))
            bar_todo = "."*((width - (len(bar_correct) + len(bar_wrong))))

            os.system("cls")

                      
            print("\u001b[38;2;60;163;112m"+bar_correct, end="")
            print("\u001b[38;2;126;126;143m"+bar_todo, end="")
            print("\u001b[38;2;176;48;92m"+bar_wrong)
            print(prev_word)
                
            
            ui = input(f"\u001b[38;2;234;169;77m{mword} > ")
            if "/" in ui:
                increase_index = False
                if ui == "/save":
                    wb.save(path)
                    saved()
                if ui == "/quit":
                    os.system('cls')
                    break
                if ui == "/repeat":
                    ask_word_again = not ask_word_again
                if ui == "/edit":
                    prev_mword = words[index-1][0]
                    prev_word = words[index-1][1]
                    prev_file_index = words[index-1][2]
                    inp = ""

                    while split_character not in inp:
                        inp = input(f"{prev_mword}{split_character}{prev_word}> ")


                    inp_list = inp.split(split_character)
                    new_word = inp_list[0]
                    new_mword = inp_list[1]
                    ws.cell(row=prev_file_index, column=2, value = str(new_word))
                    ws.cell(row=prev_file_index, column=1, value = str(new_mword))

                if ui == "/diff":
                    prev_mword = words[index-1][0]
                    prev_word = words[index-1][1]
                    difficult_words.append([prev_word, prev_mword])

                
            else:
                #right answer
                if ui == word:
                    print("\033[1A", end="")
                    print("\033[K", end="")
                    print(f"\u001b[38;2;234;169;77m{mword} > {ui}"+"\u001b[38;2;60;163;112m - correct")
                    ws.cell(row=file_index, column=3,value="correct")
                    correct_words += 1
                    prev_word = f"\u001b[38;2;234;169;77m{mword} > {ui}"+"\u001b[38;2;60;163;112m - correct"

                #wrong answer
                elif ui != word and "/" not in ui:
                    print("\033[1A", end="")
                    print("\033[K", end="")
                    print(f"\u001b[38;2;234;169;77m{mword} > {ui}"f"\u001b[38;2;176;48;92m - {word}")
                    ws.cell(row=file_index, column=3,value="wrong")
                    if ask_word_again:
                        inp = None
                        times_to_ask = 1
                        
                        while times_to_ask > 0:
                            mistake = False
                            for j in range(times_to_ask):
                                inp = input(f"\u001b[38;2;234;169;77m({j+1}/{times_to_ask})> ")
                                if inp != word:
                                    mistake = True
                                    break

                            if mistake == True:
                                times_to_ask += 2
                            else:
                                break



                    wrong_words += 1
                    prev_word = f"\u001b[38;2;234;169;77m{mword} > {ui}"f"\u001b[38;2;176;48;92m - {word}"

        #learning cycle complete
        prev_word = ""
        if ui == "quit":
            wb.save(path)
            saved()
            mode = 0
            break
        else:
        
            if len(difficult_words) != 0:
                wb.active = wb["difficult"]
                ws = wb["difficult"]

                already_difficult_words = []
                row = 1
                while True:
                    wordpair = [ws.cell(column=1, row=row).value, ws.cell(column=2, row=row).value]
                    if wordpair == [None, None]:
                        break
                    already_difficult_words.append(wordpair)
                    row += 1

                for word in difficult_words:
                    if word not in already_difficult_words:
                        ws.append(word)

                wb.active = wb[chapter]
                ws = wb[chapter]




            wb.save(path)
            saved()
            if correct_words+wrong_words == 0:
                warn(f"%correct: 0%")
            else:
                warn(f"%correct: {round((correct_words/(correct_words + wrong_words) * 100), 1)}%")
            say(f"total: {correct_words + wrong_words}")
            say(f"correct: {correct_words}")
            say(f"wrong: {wrong_words}")
            index = 0
            correct_words = 0
            wrong_words = 0
            words = []
            option = input("\u001b[38;2;234;169;77m"+"repeat(y/n)? ")
            if option == "n".lower():
                #not breaking out of loop
                mode = 0
                break
            os.system("cls")

    #add mode
    while mode == 2:
        ui = input("\u001b[38;2;60;163;112m> ")

        if split_character not in ui and "/" not in ui:
            continue
        #checks if it needs to save
        if time_to_save >= int(savetick):
            time_to_save = 0
            wb.save(path)
            saved()
        #checks for commands
        if "/" in ui:  
            
            if "/quit" in ui:
                wb.save(path)
                saved()
                mode = 0
                os.system('cls')
                break

            if ui == "/save":
                wb.save(path)
                saved()

            if "/savetick" in ui:
                say(f"current savetick: {savetick}")
                savetick = int(ui.replace("/safetick ", ""))
                settingvalues[1] = savetick
                with open(program_path+"\\"+"settings.txt", 'w', encoding='utf-8') as settings:
                    settings.writelines(settingvalues)
            
            if ui == "/replace":
                swap_word = input(f"\u001b[38;2;60;163;112mchanging: {ws.cell(row=index, column=1).value}{split_character}{ws.cell(row=index, column=2).value} > ")
                ui_list = swap_word.split(split_character)
                word = ui_list[0]
                mword = ui_list[1]
                ws.cell(row=index, column=1, value = str(word))
                ws.cell(row=index, column=2, value = str(mword))
                time_to_save +=1
                saved()

        elif "/" not in ui:
            index += 1
            time_to_save += 1
            ui_list = ui.split(split_character)
            word = ui_list[0]
            mword = ui_list[1]
            ws.cell(row=index, column=1, value = str(word))
            ws.cell(row=index, column=2, value = str(mword))
        
    if ui == "quit":
        break